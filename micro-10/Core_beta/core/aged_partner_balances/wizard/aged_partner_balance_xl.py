from odoo import api, fields, models
from openpyxl import Workbook
# from openpyxl.styles import NamedStyle as Style, Font, Alignment, Border, Side
#from openpyxl.styles import Style, Font, Alignment, Border, Side
import tempfile
import base64
import os
from datetime import date, datetime
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT
import logging
_logger = logging.getLogger(__name__)

class AccountAgedTrialBalance(models.TransientModel):
    _inherit = 'account.aged.trial.balance'

    is_excel = fields.Boolean(string='Print Excel',default=True)
    foreign_currency = fields.Boolean(string='Show Foreign Currency')
    export_format = fields.Selection([('detail','Details Report'),('summary','Summary Report')], default='detail', string='Export Format')
    file = fields.Binary('File')
    filename = fields.Char('Filename')

    @api.multi
    def action_export(self):
        if self.export_format == 'detail':
            self.action_generate_detail_xls()
        elif self.export_format == 'summary':
            self.action_generate_summary_xls()
        return {
            'name': 'Aged Partner Balance',
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=account.aged.trial.balance&id=%s&filename_field=filename&field=file&download=true&filename=%s" % (self.id, self.filename),
            'target': 'self',
        }

    @api.multi
    def action_generate_detail_xls(self):
        day1, day2, day3, day4, day5, outstanding, total_amount_list, not_due_list = [], [], [], [], [], [], [], []
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb = Workbook()
        ws = wb.active
        ws.append([self.env.user.company_id.name])
        ws.append(['Aged partner Balance Report as of ' + (datetime.today().strftime('%d/%m/%Y'))])
        ws.append([''])

        period_length = self.period_length
        column_e_header = '0' + '-' + str(period_length) + ' Days'
        period_length += period_length
        column_f_header = str(self.period_length + 1) + '-' + str(period_length) + ' Days'
        column_g_header = str(period_length + 1) + '-' + str(period_length + self.period_length) + ' Days'
        column_h_header = str(period_length + self.period_length + 1) + '-120 Days'
        foreign_currency = self.foreign_currency
        if foreign_currency:
            ws.append([
                'Partner Name',
                'Date',
                'Invoice Reference',
                'Amount (foreign currency)',
                'Currency',
                'Exchange Rate on the date',
                'Amount in ' + str(self.env.user.company_id.currency_id.name),
                'Not Due',
                column_e_header,
                column_f_header,
                column_g_header,
                column_h_header,
                '+120 days'
            ])
        else:
            ws.append([
                'Partner Name',
                'Date',
                'Invoice Reference',
                'Total Amount',
                'Not Due',
                column_e_header,
                column_f_header,
                column_g_header,
                column_h_header,
                '+120 days'
            ])
        ft = Font(size=12, bold=True)
        st = Style(font=ft)
        ws['A4'].style = st
        ws['A4'].alignment = Alignment(horizontal="center")
        ws['B4'].style = st
        ws['B4'].alignment = Alignment(horizontal="center")
        ws['C4'].style = st
        ws['C4'].alignment = Alignment(horizontal="center")
        ws['D4'].style = st
        ws['D4'].alignment = Alignment(horizontal="center")
        ws['E4'].style = st
        ws['E4'].alignment = Alignment(horizontal="center")
        ws['F4'].style = st
        ws['F4'].alignment = Alignment(horizontal="center")
        ws['G4'].style = st
        ws['G4'].alignment = Alignment(horizontal="center")
        ws['H4'].style = st
        ws['H4'].alignment = Alignment(horizontal="center")
        ws['I4'].style = st
        ws['I4'].alignment = Alignment(horizontal="center")
        ws['J4'].style = st
        ws['J4'].alignment = Alignment(horizontal="center")
        ws['K4'].style = st
        ws['K4'].alignment = Alignment(horizontal="center")
        ws['L4'].style = st
        ws['L4'].alignment = Alignment(horizontal="center")
        ws['M4'].style = st
        ws['M4'].alignment = Alignment(horizontal="center")

        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
        ws.cell(row=4, column=1).border = thin_border
        ws.cell(row=4, column=2).border = thin_border
        ws.cell(row=4, column=3).border = thin_border
        ws.cell(row=4, column=4).border = thin_border
        ws.cell(row=4, column=5).border = thin_border
        ws.cell(row=4, column=6).border = thin_border
        ws.cell(row=4, column=7).border = thin_border
        ws.cell(row=4, column=8).border = thin_border
        ws.cell(row=4, column=9).border = thin_border
        ws.cell(row=4, column=10).border = thin_border
        if foreign_currency:
            ws.cell(row=4, column=11).border = thin_border
            ws.cell(row=4, column=12).border = thin_border
            ws.cell(row=4, column=13).border = thin_border

        # Width style
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        t_count = 4

        if self.result_selection == 'customer':
            t_row_count = 4
            get_open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['out_invoice', 'in_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            user_partner_ids = self.env['res.partner'].search(['|', ('user_ids', '=', self.env.user.id), ('user_ids', '=', False)]).ids
            for invoice_obj in get_open_invoice_ids:
                if invoice_obj.partner_id.id not in user_partner_ids:
                    continue
                not_due = 0
                t_row_count += 1
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00
                if invoice_obj.date_due:
                    invoice_date = invoice_obj.date_due
                else:
                    invoice_date = invoice_obj.date_invoice
                total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(
                    invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                total_days = total_date.days
                ageing_days_1 = self.period_length
                ageing_days_2 = ageing_days_1 * 2
                ageing_days_3 = ageing_days_2 + ageing_days_1
                total_outstanding = invoice_obj.amount_total_company_signed
                outstanding.append(total_outstanding)
                if total_days < 0:
                    not_due = invoice_obj.residual
                    not_due_list.append(not_due)
                else:
                    if total_days <= ageing_days_1:
                        day_column1 = total_outstanding
                        day1.append(day_column1)
                    elif total_days <= ageing_days_2:
                        day_column2 = total_outstanding
                        day2.append(day_column2)
                    elif total_days <= ageing_days_3:
                        day_column3 = total_outstanding
                        day3.append(day_column3)
                    elif total_days <= 120:
                        upto_120_days = total_outstanding
                        day4.append(upto_120_days)
                    else:
                        other_days = total_outstanding
                        day5.append(other_days)
                total_amount = invoice_obj.amount_total
                total_amount_list.append(total_amount)
                in_date = invoice_obj.date_invoice
                in_name = invoice_obj.number
                currency = invoice_obj.currency_id.name
                exchange_rate = round(invoice_obj.amount_total_company_signed / invoice_obj.amount_total, 2)
                company_amount = invoice_obj.amount_total_company_signed
                if foreign_currency:
                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(total_amount),
                        currency,
                        '{0:,.2f}'.format(exchange_rate),
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                else:
                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws['K' + str(t_count + 1)].number_format = '0.00'
                ws['K' + str(t_count + 1)].style = st
                ws['L' + str(t_count + 1)].number_format = '0.00'
                ws['L' + str(t_count + 1)].style = st
                ws['M' + str(t_count + 1)].number_format = '0.00'
                ws['M' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1

        elif self.result_selection == 'supplier':
            t_row_count = 4
            get_open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['in_invoice', 'out_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            for invoice_obj in get_open_invoice_ids:
                not_due = 0
                t_row_count += 1
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00
                if invoice_obj.date_due:
                    invoice_date = invoice_obj.date_due
                else:
                    invoice_date = invoice_obj.date_invoice
                total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(
                    invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                total_days = total_date.days
                ageing_days_1 = self.period_length
                ageing_days_2 = ageing_days_1 * 2
                ageing_days_3 = ageing_days_2 + ageing_days_1
                total_outstanding = invoice_obj.amount_total_company_signed
                total_outstanding = -total_outstanding
                outstanding.append(total_outstanding)
                if total_days < 0:
                    not_due = invoice_obj.residual
                    not_due = -not_due
                    not_due_list.append(not_due)
                else:
                    if total_days <= ageing_days_1:
                        day_column1 = total_outstanding
                        day1.append(day_column1)
                    elif total_days <= ageing_days_2:
                        day_column2 = total_outstanding
                        day2.append(day_column2)
                    elif total_days <= ageing_days_3:
                        day_column3 = total_outstanding
                        day3.append(day_column3)
                    elif total_days <= 120:
                        upto_120_days = total_outstanding
                        day4.append(upto_120_days)
                    else:
                        other_days = total_outstanding
                        day5.append(other_days)
                total_amount = invoice_obj.amount_total
                total_amount = -total_amount
                total_amount_list.append(total_amount)
                in_date = invoice_obj.date_invoice
                in_name = invoice_obj.number
                currency = invoice_obj.currency_id.name
                exchange_rate = round(invoice_obj.amount_total_company_signed / invoice_obj.amount_total, 2)
                company_amount = invoice_obj.amount_total_company_signed
                if foreign_currency:

                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(total_amount*-1),
                        currency,
                        '{0:,.2f}'.format(exchange_rate),
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                else:
                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws['K' + str(t_count + 1)].number_format = '0.00'
                ws['K' + str(t_count + 1)].style = st
                ws['L' + str(t_count + 1)].number_format = '0.00'
                ws['L' + str(t_count + 1)].style = st
                ws['M' + str(t_count + 1)].number_format = '0.00'
                ws['M' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1
        else:
            t_row_count = 4
            get_open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['out_invoice', 'in_refund', 'in_invoice', 'out_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            for invoice_obj in get_open_invoice_ids:
                not_due = 0
                t_row_count += 1
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00
                if invoice_obj.date_due:
                    invoice_date = invoice_obj.date_due
                else:
                    invoice_date = invoice_obj.date_invoice
                total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(
                    invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                total_days = total_date.days
                ageing_days_1 = self.period_length
                ageing_days_2 = ageing_days_1 * 2
                ageing_days_3 = ageing_days_2 + ageing_days_1
                if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                    total_outstanding = invoice_obj.amount_total_company_signed
                    total_outstanding = -total_outstanding
                    outstanding.append(total_outstanding)
                else:
                    total_outstanding = invoice_obj.residual
                    outstanding.append(total_outstanding)
                if total_days < 0:
                    if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                        not_due = invoice_obj.residual
                        not_due = -not_due
                        not_due_list.append(not_due)
                    else:
                        not_due = invoice_obj.residual
                        not_due_list.append(not_due)
                else:
                    if total_days <= ageing_days_1:
                        day_column1 = total_outstanding
                        day1.append(day_column1)
                    elif total_days <= ageing_days_2:
                        day_column2 = total_outstanding
                        day2.append(day_column2)
                    elif total_days <= ageing_days_3:
                        day_column3 = total_outstanding
                        day3.append(day_column3)
                    elif total_days <= 120:
                        upto_120_days = total_outstanding
                        day4.append(upto_120_days)
                    else:
                        other_days = total_outstanding
                        day5.append(other_days)

                if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                    total_amount = invoice_obj.amount_total
                    total_amount = -total_amount
                    total_amount_list.append(total_amount)
                else:
                    total_amount = invoice_obj.amount_total
                    total_amount_list.append(total_amount)
                in_date = invoice_obj.date_invoice
                in_name = invoice_obj.number
                currency = invoice_obj.currency_id.name
                exchange_rate = round(invoice_obj.amount_total_company_signed / invoice_obj.amount_total, 2)
                company_amount = invoice_obj.amount_total_company_signed
                if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                    total_amount = total_amount*-1
                    company_amount = company_amount*-1
                if foreign_currency:
                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(total_amount),
                        currency,
                        '{0:,.2f}'.format(exchange_rate),
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                else:
                    ws.append([
                        invoice_obj.partner_id.name,
                        in_date,
                        in_name,
                        '{0:,.2f}'.format(company_amount),
                        '{0:,.2f}'.format(not_due),
                        '{0:,.2f}'.format(day_column1),
                        '{0:,.2f}'.format(day_column2),
                        '{0:,.2f}'.format(day_column3),
                        '{0:,.2f}'.format(upto_120_days),
                        '{0:,.2f}'.format(other_days)
                    ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws['K' + str(t_count + 1)].number_format = '0.00'
                ws['K' + str(t_count + 1)].style = st
                ws['L' + str(t_count + 1)].number_format = '0.00'
                ws['L' + str(t_count + 1)].style = st
                ws['M' + str(t_count + 1)].number_format = '0.00'
                ws['M' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1
        ft = Font(size=12, bold=True)
        st = Style(font=ft)
        max_row = ws.max_row
        ws.row_dimensions[1].height = 22
        ft = Font(size=14, bold=True, color='A901DB')
        st = Style(font=ft)
        ws['A1'].style = st
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:E1')
        ws.row_dimensions[2].height = 25
        ft = Font(size=18, bold=True, color='A901DB')
        st = Style(font=ft)
        ws['A2'].style = st
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:E2')

        wb.save(excelreport_path)
        excel_file_obj = open(excelreport_path, 'rb')
        bin_data = excel_file_obj.read()
        encoded_excel_data = base64.encodestring(bin_data)
        self.write({'file': encoded_excel_data, 'filename': 'Aged_Partner_Balance_Report.xlsx'})
        if excelreport_path:
            try:
                os.unlink(excelreport_path)
            except (OSError, IOError):
                _logger.error('Error when trying to remove file %s' % excelreport_path)
        return True

    @api.multi
    def action_generate_summary_xls(self):
        day1, day2, day3, day4, day5, outstanding, total_amount_list, not_due_list = [], [], [], [], [], [], [], []
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb = Workbook()
        ws = wb.active
        ws.append([self.env.user.company_id.name])
        ws.append(['Aged partner Balance Report as of ' + (datetime.today().strftime('%d/%m/%Y'))])
        ws.append([''])

        period_length = self.period_length
        column_e_header = '0' + '-' + str(period_length) + ' Days'
        period_length += period_length
        column_f_header = str(self.period_length + 1) + '-' + str(period_length) + ' Days'
        column_g_header = str(period_length + 1) + '-' + str(period_length + self.period_length) + ' Days'
        column_h_header = str(period_length + self.period_length + 1) + '-120 Days'
        ws.append([
            'Partner Name',
            'Not Due',
            column_e_header,
            column_f_header,
            column_g_header,
            column_h_header,
            '+120 days',
            'Total'
        ])
        ft = Font(size=12, bold=True)
        st = Style(font=ft)
        ws['A4'].style = st
        ws['A4'].alignment = Alignment(horizontal="center")
        ws['B4'].style = st
        ws['B4'].alignment = Alignment(horizontal="center")
        ws['C4'].style = st
        ws['C4'].alignment = Alignment(horizontal="center")
        ws['D4'].style = st
        ws['D4'].alignment = Alignment(horizontal="center")
        ws['E4'].style = st
        ws['E4'].alignment = Alignment(horizontal="center")
        ws['F4'].style = st
        ws['F4'].alignment = Alignment(horizontal="center")
        ws['G4'].style = st
        ws['G4'].alignment = Alignment(horizontal="center")
        ws['H4'].style = st
        ws['H4'].alignment = Alignment(horizontal="center")
        ws['I4'].style = st
        ws['I4'].alignment = Alignment(horizontal="center")
        ws['J4'].style = st
        ws['J4'].alignment = Alignment(horizontal="center")
        ws['K4'].style = st
        ws['K4'].alignment = Alignment(horizontal="center")
        ws['L4'].style = st
        ws['L4'].alignment = Alignment(horizontal="center")
        ws['M4'].style = st
        ws['M4'].alignment = Alignment(horizontal="center")

        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
        ws.cell(row=4, column=1).border = thin_border
        ws.cell(row=4, column=2).border = thin_border
        ws.cell(row=4, column=3).border = thin_border
        ws.cell(row=4, column=4).border = thin_border
        ws.cell(row=4, column=5).border = thin_border
        ws.cell(row=4, column=6).border = thin_border
        ws.cell(row=4, column=7).border = thin_border
        ws.cell(row=4, column=8).border = thin_border

        # Width style
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        t_count = 4

        if self.result_selection == 'customer':
            t_row_count = 4
            open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['out_invoice', 'in_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            partner_ids = list(set([invoice.partner_id.id for invoice in open_invoice_ids]))
            for partner_id in partner_ids:
                partner_obj = self.env['res.partner'].browse(partner_id)
                partner_invoice_ids = self.env['account.invoice'].search(
                [('type','in',['out_invoice','in_refund']), ('state','=','open'),
                 ('partner_id','=',partner_id), ('company_id','=',self.env.user.company_id.id),
                 ('date_invoice','<=',self.date_from)], order='id asc, number asc')
                not_due = 0
                t_row_count += 1
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days, total_amount = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00
                for invoice_obj in partner_invoice_ids:
                    if invoice_obj.date_due:
                        invoice_date = invoice_obj.date_due
                    else:
                        invoice_date = invoice_obj.date_invoice
                    total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(
                        invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                    total_days = total_date.days
                    ageing_days_1 = self.period_length
                    ageing_days_2 = ageing_days_1 * 2
                    ageing_days_3 = ageing_days_2 + ageing_days_1
                    total_outstanding = invoice_obj.amount_total_company_signed
                    outstanding.append(total_outstanding)
                    if total_days < 0:
                        not_due += invoice_obj.residual_company_signed
                    else:
                        if total_days <= ageing_days_1:
                            day_column1 += total_outstanding
                        elif total_days <= ageing_days_2:
                            day_column2 += total_outstanding
                        elif total_days <= ageing_days_3:
                            day_column3 += total_outstanding
                        elif total_days <= 120:
                            upto_120_days += total_outstanding
                        else:
                            other_days += total_outstanding
                    total_amount += total_outstanding
                not_due_list.append(not_due)
                day1.append(day_column1)
                day2.append(day_column2)
                day3.append(day_column3)
                day4.append(upto_120_days)
                day5.append(other_days)
                total_amount_list.append(total_amount)
                ws.append([
                    partner_obj.name,
                    '{0:,.2f}'.format(not_due),
                    '{0:,.2f}'.format(day_column1),
                    '{0:,.2f}'.format(day_column2),
                    '{0:,.2f}'.format(day_column3),
                    '{0:,.2f}'.format(upto_120_days),
                    '{0:,.2f}'.format(other_days),
                    '{0:,.2f}'.format(total_amount),
                ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1

        elif self.result_selection == 'supplier':
            t_row_count = 4
            open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['in_invoice', 'out_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            partner_ids = list(set([invoice.partner_id.id for invoice in open_invoice_ids]))
            for partner_id in partner_ids:
                partner_obj = self.env['res.partner'].browse(partner_id)
                partner_invoice_ids = self.env['account.invoice'].search(
                    [('type', 'in', ['in_invoice', 'out_refund']), ('state', '=', 'open'),
                     ('partner_id','=',partner_id),('company_id', '=', self.env.user.company_id.id),
                     ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days,total_amount = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,0.00
                for invoice_obj in partner_invoice_ids:
                    not_due = 0
                    t_row_count += 1
                    if invoice_obj.date_due:
                        invoice_date = invoice_obj.date_due
                    else:
                        invoice_date = invoice_obj.date_invoice
                    total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                    total_days = total_date.days
                    ageing_days_1 = self.period_length
                    ageing_days_2 = ageing_days_1 * 2
                    ageing_days_3 = ageing_days_2 + ageing_days_1
                    total_outstanding = invoice_obj.amount_total_company_signed
                    total_outstanding = -total_outstanding
                    outstanding.append(total_outstanding)

                    if total_days < 0:
                        not_due = invoice_obj.residual_company_signed
                        not_due += -not_due
                    else:
                        if total_days <= ageing_days_1:
                            day_column1 += total_outstanding
                        elif total_days <= ageing_days_2:
                            day_column2 += total_outstanding
                        elif total_days <= ageing_days_3:
                            day_column3 += total_outstanding
                        elif total_days <= 120:
                            upto_120_days += total_outstanding
                        else:
                            other_days += total_outstanding
#                     total_amount = invoice_obj.amount_total
                    total_amount += total_outstanding

                not_due_list.append(not_due)
                day1.append(day_column1)
                day2.append(day_column2)
                day3.append(day_column3)
                day4.append(upto_120_days)
                day5.append(other_days)
                total_amount_list.append(total_amount)
                ws.append([
                    partner_obj.name,
                    '{0:,.2f}'.format(not_due),
                    '{0:,.2f}'.format(day_column1),
                    '{0:,.2f}'.format(day_column2),
                    '{0:,.2f}'.format(day_column3),
                    '{0:,.2f}'.format(upto_120_days),
                    '{0:,.2f}'.format(other_days),
                    '{0:,.2f}'.format(total_amount),
                ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1
        else:
            t_row_count = 4
            open_invoice_ids = self.env['account.invoice'].search(
                [('type', 'in', ['out_invoice', 'in_refund', 'in_invoice', 'out_refund']), ('state', '=', 'open'),
                 ('company_id', '=', self.env.user.company_id.id),
                 ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
            partner_ids = list(set([invoice.partner_id.id for invoice in open_invoice_ids]))
            for partner_id in partner_ids:
                partner_obj = self.env['res.partner'].browse(partner_id)
                partner_invoice_ids = self.env['account.invoice'].search(
                    [('type', 'in', ['out_invoice', 'in_refund', 'in_invoice', 'out_refund']), ('state','=','open'),
                     ('partner_id', '=', partner_id), ('company_id', '=', self.env.user.company_id.id),
                     ('date_invoice', '<=', self.date_from)], order='id asc, number asc')
                total_outstanding, day_column1, day_column2, day_column3, upto_120_days, other_days,total_amount = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,0.00
                for invoice_obj in partner_invoice_ids:
                    not_due = 0
                    t_row_count += 1
                    if invoice_obj.date_due:
                        invoice_date = invoice_obj.date_due
                    else:
                        invoice_date = invoice_obj.date_invoice
                    total_date = datetime.strptime(self.date_from, DEFAULT_SERVER_DATE_FORMAT) - datetime.strptime(
                        invoice_date, DEFAULT_SERVER_DATE_FORMAT)
                    total_days = total_date.days
                    ageing_days_1 = self.period_length
                    ageing_days_2 = ageing_days_1 * 2
                    ageing_days_3 = ageing_days_2 + ageing_days_1
                    if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                        total_outstanding = invoice_obj.amount_total_company_signed
                        total_outstanding = -total_outstanding
                    else:
                        total_outstanding = invoice_obj.residual_company_signed
                    outstanding.append(total_outstanding)
                    if total_days < 0:
                        if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
                            not_due = invoice_obj.residual_company_signed
                            not_due += -not_due
                        else:
                            not_due += invoice_obj.residual_company_signed
                    else:
                        if total_days <= ageing_days_1:
                            day_column1 += total_outstanding
                        elif total_days <= ageing_days_2:
                            day_column2 += total_outstanding
                        elif total_days <= ageing_days_3:
                            day_column3 += total_outstanding
                        elif total_days <= 120:
                            upto_120_days += total_outstanding
                        else:
                            other_days += total_outstanding

                    if invoice_obj.type == 'in_invoice' or invoice_obj.type == 'out_refund':
#                         total_amount = total_outstanding
                        total_amount += total_outstanding
                    else:
                        total_amount += total_outstanding

                not_due_list.append(not_due)
                not_due_list.append(not_due)
                day1.append(day_column1)
                day2.append(day_column2)
                day3.append(day_column3)
                day4.append(upto_120_days)
                day5.append(other_days)
                total_amount_list.append(total_amount)
                total_amount_list.append(total_amount)

                ws.append([
                    partner_obj.name,
                    '{0:,.2f}'.format(not_due),
                    '{0:,.2f}'.format(day_column1),
                    '{0:,.2f}'.format(day_column2),
                    '{0:,.2f}'.format(day_column3),
                    '{0:,.2f}'.format(upto_120_days),
                    '{0:,.2f}'.format(other_days),
                    '{0:,.2f}'.format(total_amount)
                ])
                ft = Font(size=13)
                st = Style(font=ft)
                ws['A' + str(t_count + 1)].style = st
                ws['B' + str(t_count + 1)].style = st
                ws['C' + str(t_count + 1)].style = st
                ws['D' + str(t_count + 1)].number_format = '0.00'
                ws['D' + str(t_count + 1)].style = st
                ws['E' + str(t_count + 1)].number_format = '0.00'
                ws['E' + str(t_count + 1)].style = st
                ws['F' + str(t_count + 1)].number_format = '0.00'
                ws['F' + str(t_count + 1)].style = st
                ws['G' + str(t_count + 1)].number_format = '0.00'
                ws['G' + str(t_count + 1)].style = st
                ws['H' + str(t_count + 1)].number_format = '0.00'
                ws['H' + str(t_count + 1)].style = st
                ws['I' + str(t_count + 1)].number_format = '0.00'
                ws['I' + str(t_count + 1)].style = st
                ws['J' + str(t_count + 1)].number_format = '0.00'
                ws['J' + str(t_count + 1)].style = st
                ws.row_dimensions[t_count + 1].height = 22
                t_count += 1
        ft = Font(size=12, bold=True)
        ws.row_dimensions[1].height = 22
        ft = Font(size=14, bold=True, color='A901DB')
        st = Style(font=ft)
        ws['A1'].style = st
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:H1')
        ws.row_dimensions[2].height = 25
        ft = Font(size=18, bold=True, color='A901DB')
        st = Style(font=ft)
        ws['A2'].style = st
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:H2')

        wb.save(excelreport_path)
        excel_file_obj = open(excelreport_path, 'rb')
        bin_data = excel_file_obj.read()
        encoded_excel_data = base64.encodestring(bin_data)
        self.write({'file': encoded_excel_data, 'filename': 'Aged_Partner_Balance_Report.xlsx'})
        if excelreport_path:
            try:
                os.unlink(excelreport_path)
            except (OSError, IOError):
                _logger.error('Error when trying to remove file %s' % excelreport_path)
        return True

AccountAgedTrialBalance()
