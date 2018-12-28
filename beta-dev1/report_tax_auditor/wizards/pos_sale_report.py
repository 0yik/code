from odoo import models, fields, api, SUPERUSER_ID
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT
from openpyxl.styles import Style, Font, Alignment, Border, Side
from openpyxl import Workbook
import tempfile
import datetime
import base64
import os
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import logging
_logger = logging.getLogger(__name__)

class PosSaleReport(models.TransientModel):
    _name = 'pos.sale.report'
    _description = 'POS sales report'

    file = fields.Binary('XLS File')
    filename = fields.Char('Filename')
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    branch_id = fields.Many2one('res.branch', 'Branch')

    @api.multi
    def generate_xls(self):
        self.button_generate_xls()
        action = self.env.ref('report_tax_auditor.action_pos_sale_report').read()[0]
        action['res_id'] = self.id
        action['context'] = {'download_file': True}
        return action

    @api.multi
    def button_generate_xls(self):
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Form Penj'
        ws.sheet_view.showGridLines = False

        start_date = datetime.datetime.strptime(self.start_date, DEFAULT_SERVER_DATE_FORMAT)
        end_date = datetime.datetime.strptime(self.end_date, DEFAULT_SERVER_DATE_FORMAT)
        diff = (end_date - start_date).days

        # Journal ids
        cash_journal_id = self.env['account.journal'].search([('code','=','Cash')], limit=1)
        debit_bca_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.debit_bca_journal')
        ccard_bca_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.ccard_bca_journal')
        debit_mandiri_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.debit_mandiri_journal')
        ccard_mandiri_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.ccard_mandiri_journal')
        debit_bni_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.debit_bni_journal')
        ccard_bni_journal_id = self.env['ir.model.data'].xmlid_to_res_id('report_tax_auditor.ccard_bni_journal')

        # Font Styles
        ft_10 = Font(size=10, bold=True)
        style10 = Style(font=ft_10)
        ft_9 = Font(size=9.5)
        style9 = Style(font=ft_9)
        thin_border_all = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
        thin_border_no_bottom = Border(top=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
        thin_border_no_top = Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))

        row_count = 0
        for i in range(diff + 1):
            date1 = start_date + datetime.timedelta(i)
            date2 = date1.replace(hour=23, minute=59, second=59)
            date_now_name = date1.strftime('%d-%b-%y')
            # Details update
            ws.append([''])
            row_count += 1
            branch_name = 'SARANG OCI'
            if self.branch_id:
                branch_name += ' - ' + self.branch_id.name
            ws.append(['',branch_name])
            row_count += 1
            ws['B' + str(row_count)].style = style10
            ws.append(['','TRANSAKSI PER REGISTER'])
            row_count += 1
            ws['B' + str(row_count)].style = style10
            ws.append(['','PER',date_now_name])
            row_count += 1
            for col in ['B','C']:
                ws[col + str(row_count)].style = style10
            ws.append(['','No','No. Reg.','Pax','Penjualan','Penjualan','Penjualan','Service','Total','Tax','Grand',
                       'Discount','Terima','Selisih','Keterangan','','CASH','DEBIT','C.CARD','DEBIT','C.CARD','DEBIT','C.CARD'])
            row_count += 1
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for i in range(2, 24):
                if i==16: continue
                ws.cell(row=row_count, column=i).border = thin_border_no_bottom

            ws.append(['','','','','Makanan','Minuman','Total','Charge','Rp.','10%','Total','','Uang','','','','','BCA','BCA','MANDIRI','MANDIRI','BNI','BNI'])
            row_count += 1
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for i in range(2, 24):
                if i == 16: continue
                ws.cell(row=row_count, column=i).border = thin_border_no_top

            ws.column_dimensions['C'].width = 15

            # Pos order details
            no_count = 1
            pax_total = 0
            food_total = 0.0
            drink_total = 0.0
            food_drink_total = 0.0
            service_charge_total = 0.0
            rp_total = 0.0
            tax_total = 0.0
            grand_total = 0.0
            discount_total = 0.0
            given_amount_total = 0.0
            selisih_amount_total = 0.0
            cash_amount_total = 0.0
            debit_bca_amount_total = 0.0
            ccard_bca_amount_total = 0.0
            debit_mandiri_amount_total = 0.0
            ccard_mandiri_amount_total = 0.0
            debit_bni_amount_total = 0.0
            ccard_bni_amount_total = 0.0

            domain = [('date_order','>=',str(date1)),('date_order','<=',str(date2))]
            if self.branch_id:
                domain.append(('branch_id','=',self.branch_id.id))
            pos_orders = self.env['pos.order'].search(domain, order='id asc')
            for order in pos_orders:
                data_list = ['']
                data_list.append(no_count)
                data_list.append(order.pos_reference_no)
                data_list.append(order.customer_count)
                data_list.append('{0:,.2f}'.format(order.total_food_amount))
                data_list.append('{0:,.2f}'.format(order.total_drink_amount))
                data_list.append('{0:,.2f}'.format(order.total_food_amount+order.total_drink_amount))
                data_list.append('{0:,.2f}'.format(order.amount_service))
                data_list.append('{0:,.2f}'.format(order.total_food_amount+order.total_drink_amount+order.amount_service))
                data_list.append('{0:,.2f}'.format(order.amount_tax))
                data_list.append('{0:,.2f}'.format(order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax))
                data_list.append('{0:,.2f}'.format(0.00))
                cash_journal_amount = 0.00
                for line in order.statement_ids:
                    if line.amount > 0 and cash_journal_id and (line.journal_id.id == cash_journal_id.id):
                        cash_journal_amount += line.amount
                data_list.append('{0:,.2f}'.format(cash_journal_amount))
                data_list.append('{0:,.2f}'.format(order.amount_given_by_customer - (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax)))
                data_list.append('')
                data_list.append('')

                cash_amount = '-'
                debit_bca_amount = '-'
                ccard_bca_amount = '-'
                debit_mandiri_amount = '-'
                ccard_mandiri_amount = '-'
                debit_bni_amount = '-'
                ccard_bni_amount = '-'
                for line in order.statement_ids:
                    if line.amount > 0 and line.journal_id.id == cash_journal_id.id:
                        cash_amount = '{0:,.2f}'.format(line.amount)
                        cash_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == debit_bca_journal_id:
                        debit_bca_amount = '{0:,.2f}'.format(line.amount)
                        debit_bca_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == ccard_bca_journal_id:
                        ccard_bca_amount = '{0:,.2f}'.format(line.amount)
                        ccard_bca_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == debit_mandiri_journal_id:
                        debit_mandiri_amount = '{0:,.2f}'.format(line.amount)
                        debit_mandiri_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == ccard_mandiri_journal_id:
                        ccard_mandiri_amount = '{0:,.2f}'.format(line.amount)
                        ccard_mandiri_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == debit_bni_journal_id:
                        debit_bni_amount = '{0:,.2f}'.format(line.amount)
                        debit_bni_amount_total += line.amount
                    elif line.amount > 0 and line.journal_id.id == ccard_bni_journal_id:
                        ccard_bni_amount = '{0:,.2f}'.format(line.amount)
                        ccard_bni_amount_total += line.amount

                data_list.append(cash_amount)
                data_list.append(debit_bca_amount)
                data_list.append(ccard_bca_amount)
                data_list.append(debit_mandiri_amount)
                data_list.append(ccard_mandiri_amount)
                data_list.append(debit_bni_amount)
                data_list.append(ccard_bni_amount)
                no_count += 1
                ws.append(data_list)
                row_count += 1

                pax_total += order.customer_count
                food_total += order.total_food_amount
                drink_total += order.total_drink_amount
                food_drink_total += (order.total_food_amount + order.total_drink_amount)
                service_charge_total += order.amount_service
                rp_total += (order.total_food_amount + order.total_drink_amount + order.amount_service)
                tax_total += order.amount_tax
                grand_total += (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax)
                discount_total += 0.00
                given_amount_total += cash_journal_amount
                selisih_amount_total += (order.amount_given_by_customer - (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax))

                for col in ['E','F','G','H','I','J','K','L','M','N','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].number_format = '0.00'
                for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].style = style9
                for i in range(2, 24):
                    if i == 16: continue
                    ws.cell(row=row_count, column=i).border = thin_border_all
                for col in ['B','C','D']:
                    ws[col + str(row_count)].alignment = Alignment(horizontal="center")
                for col in ['E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].alignment = Alignment(horizontal="right")

            # Total line
            ws.append(['',date_now_name,'',pax_total,'{0:,.2f}'.format(food_total),'{0:,.2f}'.format(drink_total),
                '{0:,.2f}'.format(food_drink_total),'{0:,.2f}'.format(service_charge_total),'{0:,.2f}'.format(rp_total),
                '{0:,.2f}'.format(tax_total),'{0:,.2f}'.format(grand_total),'{0:,.2f}'.format(discount_total),
                '{0:,.2f}'.format(given_amount_total),'{0:,.2f}'.format(selisih_amount_total),'','',
                '{0:,.2f}'.format(cash_amount_total), '{0:,.2f}'.format(debit_bca_amount_total),
                '{0:,.2f}'.format(ccard_bca_amount_total), '{0:,.2f}'.format(debit_mandiri_amount_total),
                '{0:,.2f}'.format(ccard_mandiri_amount_total), '{0:,.2f}'.format(debit_bni_amount_total),
                '{0:,.2f}'.format(ccard_bni_amount_total)
            ])
            row_count += 1

            for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws[col + str(row_count)].number_format = '0.00'
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
            for col in ['B','C','D']:
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for col in ['E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].alignment = Alignment(horizontal="right")
            for i in range(2, 24):
                if i==16: continue
                ws.cell(row=row_count, column=i).border = thin_border_all
            ws.merge_cells('B'+str(row_count)+':C'+str(row_count))

        # Summary date range
        final_date_list = []
        while(1):
            date_list = []
            date_list.append(str(start_date)[:10])
            if (end_date.month > start_date.month) or (end_date.year > start_date.year):
                date_range = monthrange(start_date.year, start_date.month)
                start_date = start_date.replace(day=date_range[1])
                date_list.append(str(start_date)[:10])
                start_date += relativedelta(days=1)
                final_date_list.append(date_list)
            else:
                date_list.append(str(end_date)[:10])
                final_date_list.append(date_list)
                break

        # Indonesian month names
        month_name_dict = {1:'Januari', 2:'Februari', 3:'Maret', 4:'April', 5:'Mei', 6:'Juni', 7:'Juli', 8:'Agustus',
                           9:'September', 10:'Oktober', 11:'November', 12:'Desember'}

        # Summary data
        for date_range in final_date_list:
            start_date = datetime.datetime.strptime(date_range[0], DEFAULT_SERVER_DATE_FORMAT)
            end_date = datetime.datetime.strptime(date_range[1], DEFAULT_SERVER_DATE_FORMAT)
            diff = (end_date - start_date).days

            year_now_name = start_date.strftime('%Y')
            ws.append([''])
            ws.append([''])
            row_count += 2
            branch_name = 'SARANG OCI'
            if self.branch_id:
                branch_name += ' - ' + self.branch_id.name
            ws.append(['', branch_name])
            row_count += 1
            ws['B' + str(row_count)].style = style10
            ws.append(['', 'PENJUALAN BULAN ' + month_name_dict.get(start_date.month).upper() + ' '+year_now_name])
            row_count += 1
            ws['B' + str(row_count)].style = style10
            ws.append(
                ['', 'No', 'Tanggal', 'Guest', 'Penjualan', 'Penjualan', 'Penjualan', 'Service', 'Total', 'Tax', 'Grand',
                 'Discount', 'Terima', 'Selisih', 'Ket.', '', 'CASH', 'DEBIT', 'C.CARD', 'DEBIT', 'C.CARD', 'DEBIT', 'C.CARD'])
            row_count += 1
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for i in range(2, 24):
                if i == 16: continue
                ws.cell(row=row_count, column=i).border = thin_border_no_bottom

            ws.append(
                ['', '', '', '', '', '', '', 'Charge', 'Rp.', '10%', 'Total', '', 'Uang', '', '', '', '',
                 'BCA', 'BCA', 'MANDIRI', 'MANDIRI', 'BNI', 'BNI'])
            row_count += 1
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for i in range(2, 24):
                if i == 16: continue
                ws.cell(row=row_count, column=i).border = thin_border_no_top

            no_count = 1
            pax_total = 0
            food_total = 0.0
            drink_total = 0.0
            food_drink_total = 0.0
            service_charge_total = 0.0
            rp_total = 0.0
            tax_total = 0.0
            grand_total = 0.0
            discount_total = 0.0
            given_amount_total = 0.0
            selisih_amount_total = 0.0
            cash_amount_total = 0.0
            debit_bca_amount_total = 0.0
            ccard_bca_amount_total = 0.0
            debit_mandiri_amount_total = 0.0
            ccard_mandiri_amount_total = 0.0
            debit_bni_amount_total = 0.0
            ccard_bni_amount_total = 0.0

            for i in range(diff + 1):
                date1 = start_date + datetime.timedelta(i)
                date2 = date1.replace(hour=23, minute=59, second=59)
                date_now_name = date1.strftime('%d-%b-%y')
                # Pos order details
                pax_tot = 0
                food_tot = 0.0
                drink_tot = 0.0
                food_drink_tot = 0.0
                service_charge_tot = 0.0
                rp_tot = 0.0
                tax_tot = 0.0
                grand_tot = 0.0
                discount_tot = 0.0
                given_amount_tot = 0.0
                selisih_amount_tot = 0.0
                cash_amount_tot = 0.0
                debit_bca_amount_tot = 0.0
                ccard_bca_amount_tot = 0.0
                debit_mandiri_amount_tot = 0.0
                ccard_mandiri_amount_tot = 0.0
                debit_bni_amount_tot = 0.0
                ccard_bni_amount_tot = 0.0

                domain = [('date_order', '>=', str(date1)), ('date_order', '<=', str(date2))]
                if self.branch_id:
                    domain.append(('branch_id', '=', self.branch_id.id))
                pos_orders = self.env['pos.order'].search(domain, order='id asc')
                for order in pos_orders:
                    pax_tot += order.customer_count
                    pax_total += order.customer_count
                    food_tot += order.total_food_amount
                    food_total += order.total_food_amount
                    drink_tot += order.total_drink_amount
                    drink_total += order.total_drink_amount
                    food_drink_tot += (order.total_food_amount+order.total_drink_amount)
                    food_drink_total += (order.total_food_amount+order.total_drink_amount)
                    service_charge_tot += order.amount_service
                    service_charge_total += order.amount_service
                    rp_tot += (order.total_food_amount+order.total_drink_amount+order.amount_service)
                    rp_total += (order.total_food_amount+order.total_drink_amount+order.amount_service)
                    tax_tot += order.amount_tax
                    tax_total += order.amount_tax
                    grand_tot += (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax)
                    grand_total += (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax)
                    discount_tot += 0.00
                    discount_total += 0.00
                    given_amount_tot += order.amount_given_by_customer
                    given_amount_total += order.amount_given_by_customer
                    selisih_amount_tot += (order.amount_given_by_customer - (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax))
                    selisih_amount_total += (order.amount_given_by_customer - (order.total_food_amount+order.total_drink_amount+order.amount_service+order.amount_tax))
                    for line in order.statement_ids:
                        if line.amount > 0 and line.journal_id.id == cash_journal_id.id:
                            cash_amount_tot += line.amount
                            cash_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == debit_bca_journal_id:
                            debit_bca_amount_tot += line.amount
                            debit_bca_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == ccard_bca_journal_id:
                            ccard_bca_amount_tot += line.amount
                            ccard_bca_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == debit_mandiri_journal_id:
                            debit_mandiri_amount_tot += line.amount
                            debit_mandiri_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == ccard_mandiri_journal_id:
                            ccard_mandiri_amount_tot += line.amount
                            ccard_mandiri_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == debit_bni_journal_id:
                            debit_bni_amount_tot += line.amount
                            debit_bni_amount_total += line.amount
                        elif line.amount > 0 and line.journal_id.id == ccard_bni_journal_id:
                            ccard_bni_amount_tot += line.amount
                            ccard_bni_amount_total += line.amount

                data_list = ['']
                data_list.append(no_count)
                data_list.append(date_now_name)
                data_list.append(pax_tot)
                data_list.append('{0:,.2f}'.format(food_tot))
                data_list.append('{0:,.2f}'.format(drink_tot))
                data_list.append('{0:,.2f}'.format(food_drink_tot))
                data_list.append('{0:,.2f}'.format(service_charge_tot))
                data_list.append('{0:,.2f}'.format(rp_tot))
                data_list.append('{0:,.2f}'.format(tax_tot))
                data_list.append('{0:,.2f}'.format(grand_tot))
                data_list.append('0.00')
                data_list.append('{0:,.2f}'.format(given_amount_tot))
                data_list.append('{0:,.2f}'.format(selisih_amount_tot))
                data_list.append('')
                data_list.append('')
                data_list.append('{0:,.2f}'.format(cash_amount_tot))
                data_list.append('{0:,.2f}'.format(debit_bca_amount_tot))
                data_list.append('{0:,.2f}'.format(ccard_bca_amount_tot))
                data_list.append('{0:,.2f}'.format(debit_mandiri_amount_tot))
                data_list.append('{0:,.2f}'.format(ccard_mandiri_amount_tot))
                data_list.append('{0:,.2f}'.format(debit_bni_amount_tot))
                data_list.append('{0:,.2f}'.format(ccard_bni_amount_tot))
                no_count += 1
                ws.append(data_list)
                row_count += 1

                for col in ['E','F','G','H','I','J','K','L','M','N','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].number_format = '0.00'
                for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].style = style9
                for i in range(2, 24):
                    if i == 16: continue
                    ws.cell(row=row_count, column=i).border = thin_border_all
                for col in ['B','C','D']:
                    ws[col + str(row_count)].alignment = Alignment(horizontal="center")
                for col in ['E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                    ws[col + str(row_count)].alignment = Alignment(horizontal="right")

            # Total line
            ws.append(['','Jumlah','',pax_total,'{0:,.2f}'.format(food_total),'{0:,.2f}'.format(drink_total),
               '{0:,.2f}'.format(food_drink_total),'{0:,.2f}'.format(service_charge_total),'{0:,.2f}'.format(rp_total),
               '{0:,.2f}'.format(tax_total),'{0:,.2f}'.format(grand_total),'{0:,.2f}'.format(discount_total),
               '{0:,.2f}'.format(given_amount_total),'{0:,.2f}'.format(selisih_amount_total),'','',
               '{0:,.2f}'.format(cash_amount_total),
               '{0:,.2f}'.format(debit_bca_amount_total),'{0:,.2f}'.format(ccard_bca_amount_total),
               '{0:,.2f}'.format(debit_mandiri_amount_total),'{0:,.2f}'.format(ccard_mandiri_amount_total),
               '{0:,.2f}'.format(debit_bni_amount_total),'{0:,.2f}'.format(ccard_bni_amount_total)
            ])
            row_count += 1

            for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws[col + str(row_count)].number_format = '0.00'
            for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].style = style10
            for col in ['B','C','D']:
                ws[col + str(row_count)].alignment = Alignment(horizontal="center")
            for col in ['E','F','G','H','I','J','K','L','M','N','O','Q','R','S','T','U','V','W']:
                ws[col + str(row_count)].alignment = Alignment(horizontal="right")
            for i in range(2, 24):
                if i==16: continue
                ws.cell(row=row_count, column=i).border = thin_border_all
            ws.merge_cells('B'+str(row_count)+':C'+str(row_count))

        # Save the file
        wb.save(excelreport_path)
        excel_file_obj = open(excelreport_path, 'rb')
        bin_data = excel_file_obj.read()
        encoded_excel_data = base64.encodestring(bin_data)
        self.write({'file': encoded_excel_data, 'filename': 'pos_sale_report.xlsx'})
        if excelreport_path:
            try:
                os.unlink(excelreport_path)
            except (OSError, IOError):
                _logger.error('Error when trying to remove file %s' % excelreport_path)
        return True

PosSaleReport()