from odoo import api, fields, models
from openpyxl.styles import Style, Font, Alignment, Border, Side 
from openpyxl import Workbook
import tempfile
import base64
import os
import logging
_logger = logging.getLogger(__name__)

# core = '/home/nareshd/workspace/odoo_10_development/Core-beta/core/'
core = '/opt/odoo/odoo/addons/core/'
module_core_dir = 'Core'
basic = '/opt/odoo/odoo/addons/basic/'
module_basic_dir = 'Basic'


class db_info(models.TransientModel):
    _name = 'db.info'
    _description = 'Database module information'
    
    name = fields.Char('Name')
    file = fields.Binary('File')
    filename = fields.Char('Filename')
    
    @api.multi
    def action_export(self):
        self.action_generate_summary_xls()
        return {
            'name': 'DB Information',
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=db.info&id=%s&filename_field=filename&field=file&download=true&filename=%s" % (self.id, self.filename),
            'target': 'self',
        }
    
    @api.multi
    def action_generate_detail_xls(self):
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb = Workbook()
        ws = wb.active
        ws.append(['Database Information '])
        ws.append([''])
        ws.append(['Project Name','-',self.env.cr.dbname])
        ws.append([''])
        ws.append([
            'Module',
            'Module Name',
            'Category',
            'Summary',
            'Module Type',
            'Module Dependencies',
            'Description'
        ])
        ft = Font(size=12, bold=True)
        st = Style(font=ft)
        ws['A1'].style = st
        ws['A5'].style = st
        ws['A5'].alignment = Alignment(horizontal="center")
        ws['B5'].style = st
        ws['B5'].alignment = Alignment(horizontal="center")
        ws['C5'].style = st
        ws['C5'].alignment = Alignment(horizontal="center")
        ws['D5'].style = st
        ws['D5'].alignment = Alignment(horizontal="center")
        ws['E5'].style = st
        ws['E5'].alignment = Alignment(horizontal="center")
        ws['F5'].style = st
        ws['F5'].alignment = Alignment(horizontal="center")
        ws['G5'].style = st
        ws['G5'].alignment = Alignment(horizontal="center")
        module_obj = self.env['ir.module.module']
        module_search_obj = module_obj.search([('state','=','installed')])
        for module_id in module_search_obj:
            name_list = []
            module_name = module_id.shortdesc
            module_technical_name = module_id.name
            module_category = module_id.category_id.name
            module_summary = module_id.summary
            module_dependency = module_id.dependencies_id
            path_core = os.path.isdir(core + str(module_technical_name))
            path_basic = os.path.isdir(basic + str(module_technical_name))
            if path_core:
                module_repo = module_core_dir
            if path_basic:
                module_repo = module_basic_dir
            if not path_core and not path_basic:
                module_repo = 'Modifier'
            for dependency in module_dependency:
                name_list.append(dependency.name)
            module_dependent_name = ','.join(map(str, name_list)) 
            ws.append([
                        module_name,
                        module_technical_name,
                        module_category,
                        module_summary,
                        module_repo,
                        module_dependent_name
                     ])
            
        ws.merge_cells('A1:E1')
        wb.save(excelreport_path)
        excel_file_obj = open(excelreport_path, 'rb')
        bin_data = excel_file_obj.read()
        encoded_excel_data = base64.encodestring(bin_data)
        self.write({'file': encoded_excel_data, 'filename': str(self.env.cr.dbname)+'_db_info.xlsx'})
        if excelreport_path:
            try:
                os.unlink(excelreport_path)
            except (OSError, IOError):
                _logger.error('Error when trying to remove file %s' % excelreport_path)
        return {
            'name': 'DB Information',
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=db.info&id=%s&filename_field=filename&field=file&download=true&filename=%s" % (self.id, self.filename),
            'target': 'self',
        }

    

