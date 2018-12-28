from odoo import models, fields, api
from odoo.exceptions import UserError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT as DATE_FORMAT
from openpyxl.styles import Font, colors, Alignment
from datetime import datetime
from openpyxl import Workbook
import tempfile
import base64
import os
import logging
_logger = logging.getLogger(__name__)

class DiscrepancyReport(models.TransientModel):
    _name = 'discrepancy.report'
    _description = 'Discrepancy Report'

    @api.multi
    def _get_email_ids(self):
        for record in self:
            email_to = ''
            stock_manager_id = self.env['ir.model.data'].xmlid_to_res_id('stock.group_stock_manager')
            for user in self.env['res.users'].search([('groups_id', 'in', stock_manager_id)]):
                if user.partner_id and user.partner_id.email:
                    email_to += user.partner_id.email +','
            record.email_to = email_to[:-1] if email_to else ''

    report_file = fields.Binary('Report File')
    filename = fields.Char('File Name')
    report_name = fields.Char('Report Name')
    report_for = fields.Selection([('receiving', 'Receiving'), ('stock_count', 'Stock Count')], default='receiving', string='Selection')
    product_ids = fields.Many2many('product.product', string='Product')
    shipment_ids = fields.Many2many('shipment.reference', string='Shipment')
    email_to = fields.Char(compute='_get_email_ids', string='Email To')

    @api.onchange('report_for')
    def onchange_report_for(self):
        self.report_file = False
        self.report_name = False
        self.filename = False

    def action_generate(self):
        self.write({'report_file': False, 'filename': False, 'report_name': False})
        self.generate_excel()
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_send_mail(self):
        if not self.email_to:
            raise UserError('Inventory managers not found')
        template_id = self.env['ir.model.data'].xmlid_to_object('helaspice_receiving_import.email_template_discrepancy_report')
        if not template_id:
            raise UserError('Mail template not found.')
        vals = {
            'name': self.report_name,
            'datas': self.report_file,
            'datas_fname': self.filename,
            'res_model': 'discrepancy.report',
            'res_id': self.id,
        }
        attach_id = self.env['ir.attachment'].create(vals)
        template_id.with_context({'default_attachment_ids': [(6, 0, [attach_id.id])]}).send_mail(self.id, force_send=True)
        return True

    def generate_excel(self):
        domain = []
        if self.product_ids:
            domain += [('product_id', 'in', self.product_ids.ids)]
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb = Workbook()
        ws = wb.active
        count = 0
        ft1 = Font(size=15, bold=True, color=colors.BLUE)
        ft2 = Font(bold=True)
        if self.report_for == 'receiving':
            if self.shipment_ids:
                domain += [('shipment_id', 'in', self.shipment_ids.ids)]
            data_dict = {}
            for line in self.env['manage.incoming.batch'].search(domain, order='shipment_id').filtered(lambda x: (x.qty != x.done_qty and x.done_qty != 0) or (x.lot_id.actual_bbd or x.lot_id.actual_batch_no)):
                vals = {}
                vals['pallet_no'] = str(line.pallet_no) if line.pallet_no else ''
                vals['sh_ref'] = line.shipment_id.name if line.shipment_id else ''
                vals['po_reference'] = line.po_reference if line.po_reference else ''
                vals['item_description'] = line.product_id.name_get()[0][1]
                vals['batch_no'] = line.lot_id.batch_no
                vals['actual_batch_no'] = line.lot_id.actual_batch_no
                vals['bbd'] = str(datetime.strptime(str(line.lot_id.use_date)[:10], DATE_FORMAT).strftime('%d/%m/%Y')) if line.lot_id.use_date else ''
                vals['actual_bbd'] = str(datetime.strptime(str(line.lot_id.actual_bbd)[:10], DATE_FORMAT).strftime('%d/%m/%Y')) if line.lot_id.actual_bbd else ''
                vals['qty'] = line.qty
                vals['done_qty'] = line.done_qty
                vals['difference'] = line.done_qty - line.qty
                sh_ref = line.shipment_id.name if line.shipment_id else 'No Shipment'
                if sh_ref in data_dict:
                    data_dict[sh_ref].append(vals)
                else:
                    data_dict[sh_ref] = [vals]
            if not data_dict.keys():
                raise UserError('No data to generate report.')
            for key in data_dict.keys():
                if count > 0:
                    ws.append([''])
                    count += 1
                ws.append([('Discrepancy Report - %s') % str(key)])
                count += 1
                ws['A%s'% count].font = ft1
                ws.merge_cells(('A%s:I%s') % (count, count))
                ws['A%s'% count].alignment = Alignment(horizontal="center")
                ws.append([''])
                count += 1
                ws.append(['Pallet No', 'Shipment', 'PO Reference', 'Item Description', 'Batch No.', 'Best Before', 'Forecasted Qty', 'Received Qty', 'Difference'])
                count += 1
                for alpha in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    ws['%s%s'% (alpha, count)].font = ft2
                for data in data_dict[key]:
                    ws.append([data['pallet_no'], data['sh_ref'], data['po_reference'], data['item_description'], data['batch_no'], data['bbd'], data['qty'], data['done_qty'], data['difference']])
                    count += 1
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 52
            ws.column_dimensions['E'].width = 11
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 13
            ws.column_dimensions['I'].width = 12
            filename = 'Receiving Discrepancy Report - %s'% str(datetime.strptime(str(datetime.now().date()), DATE_FORMAT).strftime('%d/%m/%Y'))
        else:
            ws.append(['Reference', 'Count Date', 'Item Description', 'SAP Quantity', 'Warehouse Quantity', 'Difference'])
            for line in self.env['stock.count.line'].search(domain+[('count_id', '!=', False), ('state', '!=', 'close')], order='count_id desc').filtered(lambda x: x.count_qty != 0 and x.qty != x.count_qty):
                ref = line.count_id.name
                count_date = str(datetime.strptime(str(line.count_id.count_date), DATE_FORMAT).strftime('%d/%m/%Y')) if line.count_id.count_date else ''
                item_description = line.product_id.name_get()[0][1]
                qty = line.qty
                count_qty = line.count_qty
                difference = count_qty - qty
                ws.append([ref, count_date, item_description, qty, count_qty, difference])
            for alpha in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws['%s1'% alpha].font = ft2
            ws.column_dimensions['A'].width = 17
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 52
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            filename = 'Stock Count Discrepancy Report - %s'% str(datetime.strptime(str(datetime.now().date()), DATE_FORMAT).strftime('%d/%m/%Y'))
        wb.save(excelreport_path)
        excel_file_obj = open(excelreport_path, 'rb')
        bin_data = excel_file_obj.read()
        encoded_excel_data = base64.encodestring(bin_data)
        self.write({'report_file': encoded_excel_data, 'filename': '%s.xlsx'% filename, 'report_name': filename})
        if excelreport_path:
            try:
                os.unlink(excelreport_path)
            except (OSError, IOError):
                _logger.error('Error when trying to remove file %s'% excelreport_path)

DiscrepancyReport()